import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os
import logging

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("timetable_generation.log"),
        logging.StreamHandler()
    ]
)

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each)
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)
TUTORIAL_DURATION = 2  # 1 hour = 2 slots (30 mins each)
MAX_SCHEDULING_ATTEMPTS = 5000

def generate_time_slots():
    """Generate time slots for the day"""
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = (current_time + timedelta(minutes=30)).time()
        slots.append((current, next_time))
        current_time = current_time + timedelta(minutes=30)
    
    return slots

def is_elective(course):
    """Enhanced function to check if a course is an elective"""
    code = str(course['Course Code'])
    name = str(course['Course Name'])
    
    # Check for B1, B2 in course code
    if 'B1' in code or 'B2' in code:
        return True
    
    # Check for 'elective' in course name
    if 'elective' in name.lower():
        return True
        
    return False

def is_morning_break(slot):
    """Check if a time slot falls within morning break time"""
    start, end = slot
    # Morning break: 10:30-11:00
    return (time(10, 30) <= start < time(11, 0))

def is_lunch_time(slot):
    """Check if a time slot falls within lunch time range (12:30-14:30)"""
    start, end = slot
    return (time(12, 30) <= start < time(14, 30))

def load_and_clean_data():
    """Load course data from Excel or CSV file and clean it"""
    try:
        df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
        logging.info("Successfully loaded data from combined.xlsx")
    except (FileNotFoundError, Exception):
        try:
            df = pd.read_csv('combined.csv')
            logging.info("Successfully loaded data from combined.csv")
        except FileNotFoundError:
            # Try tab-delimited text file
            try:
                df = pd.read_csv('paste-7.txt', delimiter='\t')
                logging.info("Successfully loaded data from paste-7.txt")
            except FileNotFoundError:
                logging.error("Error: No valid input file found")
                exit()
        except Exception as e:
            logging.error(f"Error loading data: {str(e)}")
            exit()

    # Data cleaning steps
    for col in ['L', 'T', 'P', 'S', 'C']:
        if col in df.columns:
            df[col] = df[col].fillna(0).astype(float)
    
    required_cols = ['Department', 'Semester', 'Course Code', 'Course Name', 'L', 'T', 'P', 'Faculty', 'Classroom']
    for col in required_cols:
        if col not in df.columns:
            logging.error(f"Required column '{col}' not found")
            exit()

    # Generate missing course codes and classrooms
    for idx, row in df.iterrows():
        if pd.isna(row['Course Code']) or row['Course Code'] == "-":
            dept = str(row['Department']).strip()
            semester = str(row['Semester']).strip()
            course_name = str(row['Course Name']).strip()
            if course_name:
                words = course_name.split()
                code_part = ''.join([word[0] for word in words if word])[:3].upper()
                new_code = f"{dept[:2].upper()}{semester}{code_part}"
                df.at[idx, 'Course Code'] = new_code

        classroom = str(row['Classroom']).strip()
        if pd.isna(classroom) or classroom == "" or "Will be scheduled" in classroom:
            dept = str(row['Department']).strip()
            semester = str(row['Semester']).strip()
            df.at[idx, 'Classroom'] = f"TBD_{dept}_{semester}"

    df = df.dropna(how='all')
    df = df[(df['Department'].notna()) & (df['Department'] != "") & 
            (df['Semester'].notna()) & (df['Semester'] != "")]
    
    return df

def check_scheduling_possibility(faculty, classroom, day, start_slot, duration, professor_schedule, classroom_schedule, timetable, TIME_SLOTS):
    """Check if the given slots are available for scheduling"""
    faculty_flexible = '/' in str(faculty) or ',' in str(faculty)
    classroom_flexible = str(classroom).startswith('TBD_') or "Will be scheduled" in str(classroom)
    
    for i in range(duration):
        current_slot = start_slot + i
        if current_slot >= len(TIME_SLOTS):
            return False
            
        if timetable[day][current_slot]['type'] is not None:
            return False
            
        if is_morning_break(TIME_SLOTS[current_slot]):
            return False
            
        if not faculty_flexible and faculty in professor_schedule:
            if current_slot in professor_schedule[faculty][day]:
                return False
                
        if not classroom_flexible and classroom in classroom_schedule:
            if current_slot in classroom_schedule[classroom][day]:
                return False
    
    return True

def update_schedule(faculty, classroom, day, start_slot, duration, session_type, code, name, professor_schedule, classroom_schedule, timetable):
    """Update all schedules with the new session"""
    faculty_list = [faculty]
    if '/' in str(faculty) or ',' in str(faculty):
        faculty_list = [f.strip() for f in str(faculty).replace('/', ',').split(',')]
    
    for i in range(duration):
        for single_faculty in faculty_list:
            if single_faculty not in professor_schedule:
                professor_schedule[single_faculty] = {day: set() for day in range(len(DAYS))}
            professor_schedule[single_faculty][day].add(start_slot+i)
        
        if classroom not in classroom_schedule:
            classroom_schedule[classroom] = {day: set() for day in range(len(DAYS))}
        classroom_schedule[classroom][day].add(start_slot+i)
        
        timetable[day][start_slot+i]['type'] = session_type
        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
        timetable[day][start_slot+i]['duration'] = duration if i == 0 else 0
        timetable[day][start_slot+i]['is_first'] = (i == 0)
        timetable[day][start_slot+i]['position'] = i

def schedule_random_lunch_breaks(timetable, TIME_SLOTS):
    """Schedule lunch breaks randomly between 12:30-14:30 for each day"""
    # Find all possible lunch start slots between 12:30-14:00
    lunch_start_indices = []
    for slot_idx, slot in enumerate(TIME_SLOTS):
        start_time = slot[0]
        if time(12, 30) <= start_time < time(14, 0):
            lunch_start_indices.append(slot_idx)

    for day_idx in range(len(DAYS)):
        if lunch_start_indices:
            # Randomly select a start slot ensuring we have 2 consecutive slots
            max_start = min(len(lunch_start_indices)-1, len(TIME_SLOTS)-2)
            if max_start < 0:
                logging.warning("Could not find valid lunch slots")
                return
                
            start_idx = random.choice(lunch_start_indices[:max_start+1])
            end_idx = start_idx + 1
            
            # Mark lunch slots
            for slot_idx in range(start_idx, end_idx + 1):
                timetable[day_idx][slot_idx]['type'] = 'LUNCH'
                timetable[day_idx][slot_idx]['code'] = 'LUNCH'
                timetable[day_idx][slot_idx]['name'] = 'LUNCH BREAK'
                timetable[day_idx][slot_idx]['is_first'] = (slot_idx == start_idx)
                timetable[day_idx][slot_idx]['duration'] = 2 if slot_idx == start_idx else 0
            
            logging.info(f"Scheduled lunch on {DAYS[day_idx]} at " +
                        f"{TIME_SLOTS[start_idx][0].strftime('%H:%M')}-" +
                        f"{TIME_SLOTS[end_idx][1].strftime('%H:%M')}")
        else:
            logging.warning(f"Could not schedule lunch for {DAYS[day_idx]}")

def record_elective_as_not_scheduled(department, semester, course, summary_ws):
    """Record an elective course as not scheduled in the summary sheet"""
    code = str(course['Course Code'])
    name = str(course['Course Name'])
    faculty = str(course['Faculty'])
    classroom = str(course['Classroom'])
    
    # Record lecture component
    if course['L'] > 0:
        summary_ws.append([
            department, semester, code, name, "ELECTIVE LEC", 
            faculty, classroom, "‚ö†Ô∏è Not Scheduled", "N/A"
        ])
    
    # Record tutorial component if applicable
    if course['T'] > 0:
        summary_ws.append([
            department, semester, code, name, "ELECTIVE TUT", 
            faculty, classroom, "‚ö†Ô∏è Not Scheduled", "N/A"
        ])

def schedule_session(department, semester, course, session_type, professor_schedule, classroom_schedule, timetable, TIME_SLOTS, summary_ws, attempt_limit):
    """Schedule a specific session (lab, lecture, or tutorial)"""
    code = str(course['Course Code'])
    name = str(course['Course Name'])
    faculty = str(course['Faculty'])
    classroom = str(course['Classroom'])
    
    duration = LAB_DURATION if session_type == 'LAB' else LECTURE_DURATION if 'LEC' in session_type else TUTORIAL_DURATION
    
    scheduled = False
    attempts = 0
    
    day_load = {day: sum(1 for slot in timetable[day].values() if slot['type'] is not None) 
               for day in range(len(DAYS))}
    sorted_days = sorted(day_load.keys(), key=lambda d: day_load[d])
    
    for day in sorted_days:
        if scheduled:
            break
            
        for start_slot in range(len(TIME_SLOTS)-duration+1):
            if check_scheduling_possibility(faculty, classroom, day, start_slot, duration, 
                                         professor_schedule, classroom_schedule, timetable, TIME_SLOTS):
                update_schedule(faculty, classroom, day, start_slot, duration, session_type, 
                             code, name, professor_schedule, classroom_schedule, timetable)
                scheduled = True
                summary_ws.append([department, semester, code, name, session_type, faculty, classroom, "‚úÖ Scheduled", 
                                  f"{DAYS[day]} {TIME_SLOTS[start_slot][0].strftime('%H:%M')}"])
                break
    
    while not scheduled and attempts < attempt_limit:
        day = random.randint(0, len(DAYS)-1)
        start_slot = random.randint(0, len(TIME_SLOTS)-duration)
        
        if check_scheduling_possibility(faculty, classroom, day, start_slot, duration, 
                                     professor_schedule, classroom_schedule, timetable, TIME_SLOTS):
            update_schedule(faculty, classroom, day, start_slot, duration, session_type, 
                         code, name, professor_schedule, classroom_schedule, timetable)
            scheduled = True
            summary_ws.append([department, semester, code, name, session_type, faculty, classroom, "‚úÖ Scheduled", 
                              f"{DAYS[day]} {TIME_SLOTS[start_slot][0].strftime('%H:%M')}"])
        attempts += 1
    
    if not scheduled:
        logging.warning(f"Failed to schedule {session_type} for {code}: {name}")
        summary_ws.append([department, semester, code, name, session_type, faculty, classroom, "‚ùå Failed", "N/A"])
        
    return scheduled

def handle_lectures(department, semester, course, professor_schedule, classroom_schedule, timetable, TIME_SLOTS, summary_ws, attempt_limit):
    """Handle scheduling lectures based on L value"""
    l = int(course['L'])
    total_scheduled = 0
    failed = 0
    
    if l == 3:
        for i in range(2):
            lec_scheduled = schedule_session(
                department, semester, course, f'LEC {i+1}', 
                professor_schedule, classroom_schedule, 
                timetable, TIME_SLOTS, summary_ws, attempt_limit
            )
            if lec_scheduled:
                total_scheduled += 1
            else:
                failed += 1
    else:
        for i in range(l):
            lec_scheduled = schedule_session(
                department, semester, course, f'LEC {i+1}', 
                professor_schedule, classroom_schedule, 
                timetable, TIME_SLOTS, summary_ws, attempt_limit
            )
            if lec_scheduled:
                total_scheduled += 1
            else:
                failed += 1
    
    return total_scheduled, failed

def generate_classroom_usage_sheet(summary_ws, TIME_SLOTS, wb):
    """Generate a sheet showing when each classroom is in use"""
    classroom_usage = {}
    
    # Process scheduled sessions
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        department, semester, code, name, activity, faculty, classroom, status, time_info = row
        
        # Skip unscheduled or failed sessions
        if status != "‚úÖ Scheduled":
            continue
            
        # Parse day and time
        if "N/A" not in time_info:
            parts = time_info.split()
            if len(parts) >= 2:
                day = parts[0]
                time_str = parts[1]
                
                # Initialize classroom entry if not exists
                if classroom not in classroom_usage:
                    classroom_usage[classroom] = []
                
                # Add this usage
                classroom_usage[classroom].append({
                    'department': department,
                    'semester': semester,
                    'code': code,
                    'activity': activity,
                    'day': day,
                    'time': time_str,
                    'faculty': faculty
                })
    
    # Create classroom usage sheet
    usage_ws = wb.create_sheet(title="Classroom_Usage")
    usage_ws.append(["Classroom", "Day", "Time", "Course Code", "Activity", "Department", "Semester", "Faculty"])
    
    # Sort by classroom, then by day, then by time
    for classroom in sorted(classroom_usage.keys()):
        # Sort the activities by day and time
        activities = sorted(classroom_usage[classroom], key=lambda x: (DAYS.index(x['day']) if x['day'] in DAYS else 999, x['time']))
        
        for activity in activities:
            usage_ws.append([
                classroom,
                activity['day'],
                activity['time'],
                activity['code'],
                activity['activity'],
                activity['department'],
                activity['semester'],
                activity['faculty']
            ])
    
    # Format the worksheet
    for col_idx in range(1, 9):
        col_letter = get_column_letter(col_idx)
        usage_ws.column_dimensions[col_letter].width = 18
    
    # Apply styles to header
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in usage_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    return usage_ws

def generate_classroom_free_sheet(summary_ws, TIME_SLOTS, wb):
    """Generate a sheet showing when classrooms are free"""
    free_ws = wb.create_sheet(title="Classroom_Free_Time")
    free_ws.append(["Classroom", "Day", "Free Time Slots"])
    
    classrooms = set()
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        classrooms.add(row[6])  # Classroom is in column 7 (index 6)

    for classroom in classrooms:
        if not classroom:
            continue
            
        occupied = {day: set() for day in DAYS}
        
        for row in summary_ws.iter_rows(min_row=2, values_only=True):
            if row[6] == classroom and row[7] == "‚úÖ Scheduled":
                time_info = row[8]
                if time_info == "N/A" or '-' not in time_info:
                    continue
                    
                try:
                    day_part, time_part = time_info.split(' ', 1)
                    start_str, end_str = time_part.split('-')
                except ValueError:
                    continue
                
                # Find slot indices
                start_idx = None
                end_idx = None
                for idx, slot in enumerate(TIME_SLOTS):
                    if slot[0].strftime("%H:%M") == start_str:
                        start_idx = idx
                    if slot[1].strftime("%H:%M") == end_str:
                        end_idx = idx
                
                # Mark occupied slots
                if start_idx is not None and end_idx is not None:
                    for slot in range(start_idx, end_idx + 1):
                        occupied[day_part].add(slot)
        
        # Calculate free slots for each day
        for day in DAYS:
            all_slots = set(range(len(TIME_SLOTS)))
            free_slots = sorted(list(all_slots - occupied[day]))
            
            # Group consecutive slots
            free_periods = []
            current_start = None
            
            for slot in free_slots:
                if current_start is None:
                    current_start = slot
                elif slot != free_slots[free_slots.index(slot)-1] + 1:
                    free_periods.append((current_start, free_slots[free_slots.index(slot)-1]))
                    current_start = slot
            if current_start is not None:
                free_periods.append((current_start, free_slots[-1]))
            
            # Format time ranges
            time_ranges = []
            for start, end in free_periods:
                start_time = TIME_SLOTS[start][0].strftime("%H:%M")
                end_time = TIME_SLOTS[end][1].strftime("%H:%M")
                time_ranges.append(f"{start_time}-{end_time}")
            
            if time_ranges:
                free_ws.append([classroom, day, "\n".join(time_ranges)])
            else:
                free_ws.append([classroom, day, "No free time"])
    
    # Format the worksheet
    for col in ['A', 'B', 'C']:
        free_ws.column_dimensions[col].width = 25
        
    for row in free_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    return free_ws

def generate_all_timetables():
    """Main function to generate timetables"""
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)
    summary_ws = wb.create_sheet(title="Scheduling_Summary")
    summary_ws.append(["Department", "Semester", "Course Code", "Course Name", "Activity Type", 
                      "Faculty", "Classroom", "Scheduling Status", "Time"])
    
    professor_schedule = {}
    classroom_schedule = {}
    df = load_and_clean_data()
    
    total_courses = 0
    scheduled_courses = 0
    failed_courses = 0
    elective_courses_count = 0

    # Initialize timetables for all departments and semesters
    all_timetables = {}
    departments = df['Department'].unique()
    
    for department in departments:
        if pd.isna(department) or department == "":
            continue
            
        semester_groups = df[df['Department'] == department]['Semester'].unique()
        for semester in semester_groups:
            if pd.isna(semester) or semester == "":
                continue
                
            timetable_key = f"{department}_{semester}"
            timetable = {
                day: {
                    slot: {
                        'type': None, 
                        'code': '', 
                        'name': '', 
                        'faculty': '', 
                        'classroom': '',
                        'duration': 0,
                        'is_first': False,
                        'position': 0
                    } 
                    for slot in range(len(TIME_SLOTS))
                } 
                for day in range(len(DAYS))
            }
            
            # Pre-schedule fixed lunch breaks
            schedule_random_lunch_breaks(timetable, TIME_SLOTS)
            all_timetables[timetable_key] = timetable

    # Process all departments and semesters
    for department in departments:
        if pd.isna(department) or department == "":
            continue
            
        semester_groups = df[df['Department'] == department]['Semester'].unique()
        for semester in semester_groups:
            if pd.isna(semester) or semester == "":
                continue
                
            # Get all courses for this department/semester
            all_courses = df[(df['Department'] == department) & (df['Semester'] == str(semester))].copy()
            
            if all_courses.empty:
                continue
            
            # Separate elective and non-elective courses
            elective_courses = all_courses[all_courses.apply(is_elective, axis=1)]
            regular_courses = all_courses[~all_courses.apply(is_elective, axis=1)]
            
            # Record electives as not scheduled
            for _, elective in elective_courses.iterrows():
                record_elective_as_not_scheduled(department, semester, elective, summary_ws)
                elective_courses_count += 1
            
            timetable_key = f"{department}_{semester}"
            timetable = all_timetables[timetable_key]
            
            # Create worksheet for this department-semester
            ws_title = timetable_key[:31]  # Excel has 31 char limit for sheet names
            ws = wb.create_sheet(title=ws_title)
            
            # Priority departments get more scheduling attempts
            priority_multiplier = 1.5 if department in ['DSAI', 'ECE'] else 1
            attempt_limit = int(MAX_SCHEDULING_ATTEMPTS * priority_multiplier)
            
            # First handle courses with both labs and lectures/tutorials
            combined_courses = regular_courses[(regular_courses['P'] > 0) & ((regular_courses['L'] > 0) | (regular_courses['T'] > 0))]
            for _, course in combined_courses.iterrows():
                course_scheduled = True
                
                # Schedule lab first
                p = int(course['P'])
                if p > 0:
                    total_courses += 1
                    lab_scheduled = schedule_session(
                        department, semester, course, 'LAB', 
                        professor_schedule, classroom_schedule, 
                        timetable, TIME_SLOTS, summary_ws, attempt_limit
                    )
                    if lab_scheduled:
                        scheduled_courses += 1
                    else:
                        failed_courses += 1
                        course_scheduled = False
                
                # Schedule lectures
                l = int(course['L'])
                if l > 0:
                    # If lab failed and this is DSAI or ECE, try harder
                    current_attempt_limit = attempt_limit * 2 if not course_scheduled and department in ['DSAI', 'ECE'] else attempt_limit
                    
                    lectures_scheduled, lectures_failed = handle_lectures(
                        department, semester, course, 
                        professor_schedule, classroom_schedule, 
                        timetable, TIME_SLOTS, summary_ws, 
                        current_attempt_limit
                    )
                    
                    # If L=3, we count it as 2 courses for statistics (since we're scheduling 2 lectures)
                    if l == 3:
                        total_courses += 2
                    else:
                        total_courses += l
                        
                    scheduled_courses += lectures_scheduled
                    failed_courses += lectures_failed
                
                # Schedule tutorials
                t = int(course['T'])
                for tutorial_idx in range(t):
                    total_courses += 1
                    # If lab failed and this is DSAI or ECE, try harder
                    current_attempt_limit = attempt_limit * 2 if not course_scheduled and department in ['DSAI', 'ECE'] else attempt_limit
                    
                    tut_scheduled = schedule_session(
                        department, semester, course, f'TUT {tutorial_idx+1}', 
                        professor_schedule, classroom_schedule, 
                        timetable, TIME_SLOTS, summary_ws, current_attempt_limit
                    )
                    if tut_scheduled:
                        scheduled_courses += 1
                    else:
                        failed_courses += 1
            
            # Process remaining labs
            lab_courses = regular_courses[(regular_courses['P'] > 0) & ~((regular_courses['L'] > 0) | (regular_courses['T'] > 0))]
            for _, course in lab_courses.iterrows():
                total_courses += 1
                lab_scheduled = schedule_session(
                    department, semester, course, 'LAB', 
                    professor_schedule, classroom_schedule, 
                    timetable, TIME_SLOTS, summary_ws, attempt_limit
                )
                if lab_scheduled:
                    scheduled_courses += 1
                else:
                    failed_courses += 1
            
            # Process remaining lectures and tutorials
            other_courses = regular_courses[regular_courses['P'] == 0]
            for _, course in other_courses.iterrows():
                l = int(course['L'])
                if l > 0:
                    lectures_scheduled, lectures_failed = handle_lectures(
                        department, semester, course, 
                        professor_schedule, classroom_schedule, 
                        timetable, TIME_SLOTS, summary_ws, 
                        attempt_limit
                    )
                    
                    # If L=3, we count it as 2 courses for statistics (since we're scheduling 2 lectures)
                    if l == 3:
                        total_courses += 2
                    else:
                        total_courses += l
                        
                    scheduled_courses += lectures_scheduled
                    failed_courses += lectures_failed
                
                # Schedule tutorials
                t = int(course['T'])
                for tutorial_idx in range(t):
                    total_courses += 1
                    tut_scheduled = schedule_session(
                        department, semester, course, f'TUT {tutorial_idx+1}', 
                        professor_schedule, classroom_schedule, 
                        timetable, TIME_SLOTS, summary_ws, attempt_limit
                    )
                    if tut_scheduled:
                        scheduled_courses += 1
                    else:
                        failed_courses += 1
            
            # Write timetable to worksheet
            header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
            ws.append(header)
            
            # Apply header formatting
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Define fill colors for different session types
            lec_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Lavender
            lab_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")  # Pale Green
            tut_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Misty Rose
            break_fill = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid") # Light Gray
            lunch_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid") # Light Salmon
            conflict_fill = PatternFill(start_color="FF4500", end_color="FF4500", fill_type="solid") # Tomato
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Write timetable data
            for day_idx, day in enumerate(DAYS):
                row_num = day_idx + 2  # +1 for header, +1 because rows start at 1
                ws.append([day])
                
                # First, mark all occupied cells
                occupied_cells = [False] * len(TIME_SLOTS)
                
                # Mark break times
                for slot_idx in range(len(TIME_SLOTS)):
                    if is_morning_break(TIME_SLOTS[slot_idx]):
                        occupied_cells[slot_idx] = True
                
                # Track which cells need to be merged and their merge ranges
                merges = {}  # key: start slot index, value: (end slot index, activity details)
                
                # First pass - identify merges 
                for slot_idx in range(len(TIME_SLOTS)):
                    slot_info = timetable[day_idx][slot_idx]
                    
                    # If this is the first slot of a multi-slot activity
                    if slot_info['is_first'] and slot_info['duration'] > 1:
                        end_slot = slot_idx + slot_info['duration'] - 1
                        # Check if any of these slots are already marked as occupied
                        conflict = False
                        for i in range(slot_idx, end_slot + 1):
                            if i >= len(TIME_SLOTS) or occupied_cells[i]:
                                conflict = True
                                break
                                
                        if not conflict:
                            # Mark all these slots as occupied
                            for i in range(slot_idx, end_slot + 1):
                                occupied_cells[i] = True
                            # Store the merge information
                            merges[slot_idx] = (end_slot, {
                                'type': slot_info['type'],
                                'code': slot_info['code'],
                                'name': slot_info['name'],
                                'faculty': slot_info['faculty'],
                                'classroom': slot_info['classroom']
                            })
                        else:
                            # Mark only this cell as occupied - it's a conflict
                            occupied_cells[slot_idx] = True
                    
                    # For single-slot activities (like break times or conflict indicators)
                    elif slot_info['type'] is not None and not occupied_cells[slot_idx]:
                        occupied_cells[slot_idx] = True
                
                # Second pass - write cells and perform merges
                for slot_idx in range(len(TIME_SLOTS)):
                    cell_content = ''
                    cell_fill = None
                    
                    # First priority: morning breaks
                    if is_morning_break(TIME_SLOTS[slot_idx]):
                        cell_content = "MORNING BREAK"
                        cell_fill = break_fill
                    
                    # Second priority: merge start points
                    elif slot_idx in merges:
                        end_slot, activity = merges[slot_idx]
                        activity_type = activity['type']
                        
                        # Set fill color based on activity type
                        if activity_type == 'LUNCH':
                            cell_fill = lunch_fill
                            cell_content = "üç± LUNCH BREAK"
                        elif 'LEC' in activity_type:
                            cell_fill = lec_fill
                            cell_content = f"{activity['code']} {activity_type}\n{activity['name']}\n{activity['faculty']}\n{activity['classroom']}"
                        elif activity_type == 'LAB':
                            cell_fill = lab_fill
                            cell_content = f"{activity['code']} {activity_type}\n{activity['name']}\n{activity['faculty']}\n{activity['classroom']}"
                        else:  # TUT
                            cell_fill = tut_fill
                            cell_content = f"{activity['code']} {activity_type}\n{activity['name']}\n{activity['faculty']}\n{activity['classroom']}"
                        
                        # Merge cells
                        start_col = get_column_letter(slot_idx + 2)  # +1 for day column, +1 for 1-based index
                        end_col = get_column_letter(end_slot + 2)
                        try:
                            ws.merge_cells(f"{start_col}{row_num}:{end_col}{row_num}")
                        except Exception as e:
                            logging.warning(f"Failed to merge cells for {activity_type} on {day}: {str(e)}")
                    
                    # Third priority: cells that are part of a merged range (skip them)
                    elif any(slot_idx > start and slot_idx <= end for start, (end, _) in merges.items()):
                        continue
                    
                    # Fourth priority: individual activities or conflict markers
                    elif timetable[day_idx][slot_idx]['type'] is not None:
                        code = timetable[day_idx][slot_idx]['code']
                        activity_type = timetable[day_idx][slot_idx]['type']
                        
                        # Check if this should be a merged cell but couldn't be merged
                        if timetable[day_idx][slot_idx]['is_first'] and timetable[day_idx][slot_idx]['duration'] > 1:
                            cell_content = f"üõë {code} {activity_type} - CONFLICT"
                            cell_fill = conflict_fill
                        else:
                            # Regular single-slot activity
                            name = timetable[day_idx][slot_idx]['name']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            
                            if activity_type == 'LUNCH':
                                cell_content = "üç± LUNCH BREAK"
                                cell_fill = lunch_fill
                            else:
                                cell_content = f"‚úèÔ∏è {code} {activity_type}\n{name}\n{faculty}\n{classroom}" if 'LEC' in activity_type else f"üß™ {code} {activity_type}\n{name}\n{faculty}\n{classroom}" if activity_type == 'LAB' else f"üìò {code} {activity_type}\n{name}\n{faculty}\n{classroom}"
                                
                                # Set fill color based on activity type
                                if 'LEC' in activity_type:
                                    cell_fill = lec_fill
                                elif activity_type == 'LAB':
                                    cell_fill = lab_fill
                                else:  # TUT
                                    cell_fill = tut_fill
                    
                    # Write the cell content and apply formatting
                    if cell_content:
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_content)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Adjust column widths and row heights
            for col_idx in range(1, len(TIME_SLOTS)+2):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 18  # Slightly wider columns for better readability
            
            for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                ws.row_dimensions[row[0].row].height = 80  # Taller rows for better readability
    
    # Format summary worksheet
    for col_idx in range(1, 10):  # One more column for time
        col_letter = get_column_letter(col_idx)
        summary_ws.column_dimensions[col_letter].width = 20
    
    # Apply styles to summary sheet
    for row in summary_ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    # Generate classroom usage sheet
    generate_classroom_usage_sheet(summary_ws, TIME_SLOTS, wb)
    
    # generate_classroom_free_sheet(summary_ws, TIME_SLOTS, wb)
    # Add summary statistics
    stats_ws = wb.create_sheet(title="Statistics", index=0)
    stats_ws.append(["Timetable Generation Statistics"])
    stats_ws.append(["Total courses processed:", total_courses])
    stats_ws.append(["Elective courses skipped:", elective_courses_count])
    stats_ws.append(["Successfully scheduled:", scheduled_courses])
    stats_ws.append(["Failed to schedule:", failed_courses])
    stats_ws.append(["Success rate:", f"{scheduled_courses/total_courses*100:.2f}%" if total_courses > 0 else "N/A"])
    
    # Department-wise statistics
    stats_ws.append([])
    stats_ws.append(["Department-wise Statistics:"])
    stats_ws.append(["Department", "‚úÖ Scheduled", "‚ùå Failed", "Success Rate"])
    
    # Calculate department-wise statistics
    dept_stats = {}
    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        dept = row[0]
        status = row[7]
        if dept not in dept_stats:
            dept_stats[dept] = {'scheduled': 0, 'failed': 0}
        
        if status == "‚úÖ Scheduled":
            dept_stats[dept]['scheduled'] += 1
        elif status == "‚ùå Failed":
            dept_stats[dept]['failed'] += 1
    
    # Add department statistics to worksheet
    for dept, stats in dept_stats.items():
        total = stats['scheduled'] + stats['failed']
        success_rate = (stats['scheduled'] / total * 100) if total > 0 else 0
        stats_ws.append([dept, stats['scheduled'], stats['failed'], f"{success_rate:.2f}%"])
    
    # Apply formatting to stats sheet
    stats_ws.column_dimensions['A'].width = 25
    stats_ws.column_dimensions['B'].width = 15
    stats_ws.row_dimensions[1].height = 30
    
    stats_ws['A1'].font = Font(bold=True, size=14)
    stats_ws.merge_cells('A1:B1')
    stats_ws['A1'].alignment = Alignment(horizontal='center')
    
    # Save workbook
    output_file = "timetables_no_electives.xlsx"
    try:
        wb.save(output_file)
        logging.info(f"Timetables saved to {output_file}")
    except PermissionError:
        alt_file = f"timetables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(alt_file)
        logging.warning(f"Saved to {alt_file} due to permission error")

if __name__ == "__main__":
    generate_all_timetables()